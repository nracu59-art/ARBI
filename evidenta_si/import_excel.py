"""Import initial data from Excel into SQLite database."""
import sys
import os
import openpyxl
from datetime import datetime

sys.path.insert(0, os.path.dirname(__file__))
from models import get_db, init_db

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..',
    '..', '.claude', 'uploads',
    '615ca1d7-c3fd-4dc7-b10e-f6c4276feed3',
    '1732ca79-ARBI__Baze_de_date.xlsx')


def parse_date(val):
    if val is None or val == '-' or val == '':
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    return str(val).strip() if str(val).strip() != '-' else None


def import_from_excel(excel_path=None):
    init_db()
    path = excel_path or EXCEL_PATH
    wb = openpyxl.load_workbook(path)
    conn = get_db()

    # Insert all unique systems from Analiza sheet
    ws_analiza = wb['Analiza']
    sisteme_inserted = {}
    for row in ws_analiza.iter_rows(min_row=4, values_only=True):
        denumire, nr_utilizatori = row[0], row[1]
        if denumire and denumire != 'Total' and isinstance(denumire, str):
            denumire = denumire.strip()
            cursor = conn.execute(
                "INSERT OR IGNORE INTO sisteme (denumire, statut) VALUES (?, 'activ')",
                (denumire,)
            )
            if cursor.lastrowid:
                sisteme_inserted[denumire] = cursor.lastrowid
            else:
                r = conn.execute("SELECT id FROM sisteme WHERE denumire = ?", (denumire,)).fetchone()
                if r:
                    sisteme_inserted[denumire] = r['id']

    conn.commit()

    # Load all system IDs
    all_sisteme = {r['denumire']: r['id'] for r in conn.execute("SELECT id, denumire FROM sisteme")}

    # Insert users from Registrul utilizatorilor sheet
    ws_users = wb['Registrul utilizatorilor']
    inserted = 0
    skipped = 0
    for row in ws_users.iter_rows(min_row=2, values_only=True):
        login, nume, din_data, pina_la, registru, statut = row
        if not login or not registru:
            skipped += 1
            continue
        registru = str(registru).strip()
        if registru not in all_sisteme:
            # Insert missing system
            cursor = conn.execute(
                "INSERT OR IGNORE INTO sisteme (denumire, statut) VALUES (?, 'activ')",
                (registru,)
            )
            conn.commit()
            r = conn.execute("SELECT id FROM sisteme WHERE denumire = ?", (registru,)).fetchone()
            all_sisteme[registru] = r['id']

        sistem_id = all_sisteme[registru]
        conn.execute(
            "INSERT INTO utilizatori (login, nume, sistem_id, din_data, pina_la, statut) VALUES (?,?,?,?,?,?)",
            (
                str(login).strip(),
                str(nume).strip() if nume else '',
                sistem_id,
                parse_date(din_data),
                parse_date(pina_la),
                str(statut).strip() if statut else 'activ',
            )
        )
        inserted += 1

    conn.commit()
    conn.close()
    print(f"Import complet: {len(all_sisteme)} sisteme, {inserted} utilizatori importati, {skipped} randuri sarite.")


if __name__ == '__main__':
    path = sys.argv[1] if len(sys.argv) > 1 else None
    import_from_excel(path)
