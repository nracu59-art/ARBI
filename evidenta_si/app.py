from flask import Flask, render_template, request, redirect, url_for, flash, send_file
import io
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from models import get_db, init_db

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'arbi-evidenta-si-2025')


# ── Helpers ──────────────────────────────────────────────────────────────────

def paginate(query_rows, page, per_page=20):
    total = len(query_rows)
    start = (page - 1) * per_page
    items = query_rows[start:start + per_page]
    total_pages = (total + per_page - 1) // per_page
    return items, total, total_pages


# ── Dashboard ─────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    conn = get_db()
    stats = {
        'total_sisteme': conn.execute("SELECT COUNT(*) FROM sisteme").fetchone()[0],
        'sisteme_active': conn.execute("SELECT COUNT(*) FROM sisteme WHERE statut='activ'").fetchone()[0],
        'total_utilizatori': conn.execute("SELECT COUNT(*) FROM utilizatori").fetchone()[0],
        'utilizatori_activi': conn.execute("SELECT COUNT(*) FROM utilizatori WHERE statut='activ'").fetchone()[0],
    }
    top_sisteme = conn.execute("""
        SELECT s.denumire, COUNT(u.id) as nr_utilizatori
        FROM sisteme s
        LEFT JOIN utilizatori u ON u.sistem_id = s.id
        GROUP BY s.id
        ORDER BY nr_utilizatori DESC
        LIMIT 10
    """).fetchall()
    conn.close()
    return render_template('index.html', stats=stats, top_sisteme=top_sisteme)


# ── Sisteme informaționale ────────────────────────────────────────────────────

@app.route('/sisteme')
def sisteme_list():
    q = request.args.get('q', '').strip()
    statut = request.args.get('statut', '')
    page = max(1, int(request.args.get('page', 1)))
    conn = get_db()
    sql = """
        SELECT s.*, COUNT(u.id) as nr_utilizatori
        FROM sisteme s
        LEFT JOIN utilizatori u ON u.sistem_id = s.id
        WHERE 1=1
    """
    params = []
    if q:
        sql += " AND (s.denumire LIKE ? OR s.furnizor LIKE ? OR s.descriere LIKE ?)"
        params += [f'%{q}%', f'%{q}%', f'%{q}%']
    if statut:
        sql += " AND s.statut = ?"
        params.append(statut)
    sql += " GROUP BY s.id ORDER BY s.denumire"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    items, total, total_pages = paginate(rows, page)
    return render_template('sisteme_list.html', sisteme=items, total=total,
                           page=page, total_pages=total_pages, q=q, statut=statut)


@app.route('/sisteme/adauga', methods=['GET', 'POST'])
def sistem_adauga():
    if request.method == 'POST':
        denumire = request.form.get('denumire', '').strip()
        if not denumire:
            flash('Denumirea este obligatorie.', 'danger')
            return render_template('sistem_form.html', sistem=None)
        conn = get_db()
        existing = conn.execute("SELECT id FROM sisteme WHERE denumire = ?", (denumire,)).fetchone()
        if existing:
            conn.close()
            flash('Există deja un sistem cu această denumire.', 'warning')
            return render_template('sistem_form.html', sistem=request.form)
        conn.execute(
            "INSERT INTO sisteme (denumire, furnizor, tip_acces, descriere, data_inregistrare, statut, observatii) VALUES (?,?,?,?,?,?,?)",
            (denumire, request.form.get('furnizor'), request.form.get('tip_acces'),
             request.form.get('descriere'), request.form.get('data_inregistrare'),
             request.form.get('statut', 'activ'), request.form.get('observatii'))
        )
        conn.commit()
        conn.close()
        flash('Sistemul informațional a fost adăugat.', 'success')
        return redirect(url_for('sisteme_list'))
    return render_template('sistem_form.html', sistem=None)


@app.route('/sisteme/<int:sistem_id>/editeaza', methods=['GET', 'POST'])
def sistem_editeaza(sistem_id):
    conn = get_db()
    sistem = conn.execute("SELECT * FROM sisteme WHERE id = ?", (sistem_id,)).fetchone()
    if not sistem:
        conn.close()
        flash('Sistemul nu a fost găsit.', 'danger')
        return redirect(url_for('sisteme_list'))
    if request.method == 'POST':
        denumire = request.form.get('denumire', '').strip()
        if not denumire:
            flash('Denumirea este obligatorie.', 'danger')
            conn.close()
            return render_template('sistem_form.html', sistem=sistem)
        conn.execute(
            "UPDATE sisteme SET denumire=?, furnizor=?, tip_acces=?, descriere=?, data_inregistrare=?, statut=?, observatii=? WHERE id=?",
            (denumire, request.form.get('furnizor'), request.form.get('tip_acces'),
             request.form.get('descriere'), request.form.get('data_inregistrare'),
             request.form.get('statut', 'activ'), request.form.get('observatii'), sistem_id)
        )
        conn.commit()
        conn.close()
        flash('Sistemul informațional a fost actualizat.', 'success')
        return redirect(url_for('sisteme_list'))
    conn.close()
    return render_template('sistem_form.html', sistem=sistem)


@app.route('/sisteme/<int:sistem_id>/sterge', methods=['POST'])
def sistem_sterge(sistem_id):
    conn = get_db()
    conn.execute("DELETE FROM sisteme WHERE id = ?", (sistem_id,))
    conn.commit()
    conn.close()
    flash('Sistemul informațional a fost șters.', 'success')
    return redirect(url_for('sisteme_list'))


@app.route('/sisteme/<int:sistem_id>')
def sistem_detalii(sistem_id):
    conn = get_db()
    sistem = conn.execute("SELECT * FROM sisteme WHERE id = ?", (sistem_id,)).fetchone()
    if not sistem:
        conn.close()
        flash('Sistemul nu a fost găsit.', 'danger')
        return redirect(url_for('sisteme_list'))
    utilizatori = conn.execute(
        "SELECT * FROM utilizatori WHERE sistem_id = ? ORDER BY nume, login",
        (sistem_id,)
    ).fetchall()
    conn.close()
    return render_template('sistem_detalii.html', sistem=sistem, utilizatori=utilizatori)


# ── Utilizatori ───────────────────────────────────────────────────────────────

@app.route('/utilizatori')
def utilizatori_list():
    q = request.args.get('q', '').strip()
    sistem_id = request.args.get('sistem_id', '')
    statut = request.args.get('statut', '')
    page = max(1, int(request.args.get('page', 1)))
    conn = get_db()
    sql = """
        SELECT u.*, s.denumire as sistem_denumire
        FROM utilizatori u
        JOIN sisteme s ON s.id = u.sistem_id
        WHERE 1=1
    """
    params = []
    if q:
        sql += " AND (u.login LIKE ? OR u.nume LIKE ?)"
        params += [f'%{q}%', f'%{q}%']
    if sistem_id:
        sql += " AND u.sistem_id = ?"
        params.append(sistem_id)
    if statut:
        sql += " AND u.statut = ?"
        params.append(statut)
    sql += " ORDER BY u.nume, u.login"
    rows = conn.execute(sql, params).fetchall()
    sisteme = conn.execute("SELECT id, denumire FROM sisteme ORDER BY denumire").fetchall()
    conn.close()
    items, total, total_pages = paginate(rows, page)
    return render_template('utilizatori_list.html', utilizatori=items, total=total,
                           page=page, total_pages=total_pages, q=q,
                           sistem_id=sistem_id, statut=statut, sisteme=sisteme)


@app.route('/utilizatori/adauga', methods=['GET', 'POST'])
def utilizator_adauga():
    conn = get_db()
    sisteme = conn.execute("SELECT id, denumire FROM sisteme WHERE statut='activ' ORDER BY denumire").fetchall()
    if request.method == 'POST':
        login = request.form.get('login', '').strip()
        nume = request.form.get('nume', '').strip()
        sistem_id = request.form.get('sistem_id', '').strip()
        if not login or not nume or not sistem_id:
            flash('Login, Nume și Sistemul sunt obligatorii.', 'danger')
            conn.close()
            return render_template('utilizator_form.html', utilizator=request.form, sisteme=sisteme)
        conn.execute(
            "INSERT INTO utilizatori (login, nume, sistem_id, din_data, pina_la, statut, observatii) VALUES (?,?,?,?,?,?,?)",
            (login, nume, int(sistem_id),
             request.form.get('din_data') or None,
             request.form.get('pina_la') or None,
             request.form.get('statut', 'activ'),
             request.form.get('observatii'))
        )
        conn.commit()
        conn.close()
        flash('Utilizatorul a fost adăugat.', 'success')
        return redirect(url_for('utilizatori_list'))
    conn.close()
    return render_template('utilizator_form.html', utilizator=None, sisteme=sisteme)


@app.route('/utilizatori/<int:util_id>/editeaza', methods=['GET', 'POST'])
def utilizator_editeaza(util_id):
    conn = get_db()
    util = conn.execute("SELECT * FROM utilizatori WHERE id = ?", (util_id,)).fetchone()
    if not util:
        conn.close()
        flash('Utilizatorul nu a fost găsit.', 'danger')
        return redirect(url_for('utilizatori_list'))
    sisteme = conn.execute("SELECT id, denumire FROM sisteme ORDER BY denumire").fetchall()
    if request.method == 'POST':
        login = request.form.get('login', '').strip()
        nume = request.form.get('nume', '').strip()
        sistem_id = request.form.get('sistem_id', '').strip()
        if not login or not nume or not sistem_id:
            flash('Login, Nume și Sistemul sunt obligatorii.', 'danger')
            conn.close()
            return render_template('utilizator_form.html', utilizator=util, sisteme=sisteme)
        conn.execute(
            "UPDATE utilizatori SET login=?, nume=?, sistem_id=?, din_data=?, pina_la=?, statut=?, observatii=? WHERE id=?",
            (login, nume, int(sistem_id),
             request.form.get('din_data') or None,
             request.form.get('pina_la') or None,
             request.form.get('statut', 'activ'),
             request.form.get('observatii'), util_id)
        )
        conn.commit()
        conn.close()
        flash('Utilizatorul a fost actualizat.', 'success')
        return redirect(url_for('utilizatori_list'))
    conn.close()
    return render_template('utilizator_form.html', utilizator=util, sisteme=sisteme)


@app.route('/utilizatori/<int:util_id>/sterge', methods=['POST'])
def utilizator_sterge(util_id):
    conn = get_db()
    conn.execute("DELETE FROM utilizatori WHERE id = ?", (util_id,))
    conn.commit()
    conn.close()
    flash('Utilizatorul a fost șters.', 'success')
    return redirect(url_for('utilizatori_list'))


# ── Import / Export ───────────────────────────────────────────────────────────

@app.route('/import', methods=['GET', 'POST'])
def import_excel():
    if request.method == 'POST':
        f = request.files.get('fisier')
        if not f or not f.filename.endswith('.xlsx'):
            flash('Selectați un fișier .xlsx valid.', 'danger')
            return render_template('import.html')
        tmp_path = '/tmp/import_si.xlsx'
        f.save(tmp_path)
        try:
            from import_excel import import_from_excel
            import_from_excel(tmp_path)
            flash('Import realizat cu succes.', 'success')
        except Exception as e:
            flash(f'Eroare la import: {e}', 'danger')
        return redirect(url_for('index'))
    return render_template('import.html')


@app.route('/export/utilizatori')
def export_utilizatori():
    q = request.args.get('q', '').strip()
    sistem_id = request.args.get('sistem_id', '')
    statut = request.args.get('statut', '')
    conn = get_db()
    sql = """
        SELECT u.login, u.nume, s.denumire as sistem, u.din_data, u.pina_la, u.statut, u.observatii
        FROM utilizatori u
        JOIN sisteme s ON s.id = u.sistem_id
        WHERE 1=1
    """
    params = []
    if q:
        sql += " AND (u.login LIKE ? OR u.nume LIKE ?)"
        params += [f'%{q}%', f'%{q}%']
    if sistem_id:
        sql += " AND u.sistem_id = ?"
        params.append(sistem_id)
    if statut:
        sql += " AND u.statut = ?"
        params.append(statut)
    sql += " ORDER BY u.nume, u.login"
    rows = conn.execute(sql, params).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registrul utilizatorilor"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ['Login', 'Numele deținătorului', 'Sistemul', 'Din data', 'Pînă', 'Statut', 'Observații']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    col_widths = [15, 30, 55, 15, 15, 10, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"registrul_utilizatorilor_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/export/sisteme')
def export_sisteme():
    conn = get_db()
    rows = conn.execute("""
        SELECT s.denumire, s.furnizor, s.tip_acces, s.descriere,
               s.data_inregistrare, s.statut, COUNT(u.id) as nr_utilizatori, s.observatii
        FROM sisteme s
        LEFT JOIN utilizatori u ON u.sistem_id = s.id
        GROUP BY s.id
        ORDER BY s.denumire
    """).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sisteme Informationale"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ['Denumire', 'Furnizor', 'Tip acces', 'Descriere', 'Data înregistrării', 'Statut', 'Nr. utilizatori', 'Observații']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    col_widths = [55, 25, 20, 40, 20, 10, 15, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"sisteme_informationale_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='0.0.0.0', port=5001)
