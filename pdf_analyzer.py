"""
Analizează PDF-urile atașate hotărârilor penale zilnice și caută mențiuni de
confiscare/sechestru. Generează un raport Excel și îl trimite pe Telegram.
"""
import asyncio
import io
import json
import logging
import os
import sys
from datetime import date, timedelta
from pathlib import Path

import httpx
import pdfplumber
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from scraper import scrape_decisions

load_dotenv()
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

TELEGRAM_API_MSG = "https://api.telegram.org/bot{token}/sendMessage"
TELEGRAM_API_DOC = "https://api.telegram.org/bot{token}/sendDocument"

_CONFIG_PATH = Path(__file__).parent / "config" / "keywords.json"


def _load_keywords() -> list[str]:
    """Încarcă keywords din config/keywords.json. Fallback la lista hardcodată."""
    try:
        with open(_CONFIG_PATH, encoding="utf-8") as f:
            cfg = json.load(f)
        kws = cfg.get("case_filter", [])
        if kws:
            logger.info(f"Keywords încărcate din {_CONFIG_PATH}: {len(kws)} cuvinte")
            return kws
    except Exception as exc:
        logger.warning(f"Nu s-a putut citi config/keywords.json: {exc} — se folosesc keywords implicite")
    return [
        "confiscare", "confiscării", "confiscat", "confiscate",
        "în folosul statului", "sechestru", "sechestrat", "sechestrate",
        "se ridică sechestrul", "menține sechestrul", "menținerea sechestrului",
        "trecut cu titlu gratuit", "trecerea cu titlu gratuit",
    ]


KEYWORDS = _load_keywords()


CONTEXT_CHARS = 400  # caractere extrase în jurul cuvântului cheie

STATUS_MATCH = "Mențiuni găsite"
STATUS_NO_MATCH = "Fără mențiuni confiscare"
STATUS_UNAVAILABLE = "PDF indisponibil"

_MONTHS_RO = [
    "", "ianuarie", "februarie", "martie", "aprilie", "mai", "iunie",
    "iulie", "august", "septembrie", "octombrie", "noiembrie", "decembrie",
]


def _require(name: str) -> str:
    value = os.getenv(name)
    if not value:
        logger.error(f"Missing required environment variable: {name}")
        sys.exit(1)
    return value


def _day_ro(d: date) -> str:
    return f"{d.day} {_MONTHS_RO[d.month]} {d.year}"


# ─────────────────────────────────────────────────────────────────────────────
# PDF download & text extraction
# ─────────────────────────────────────────────────────────────────────────────

async def download_pdf(url: str) -> bytes | None:
    if not url:
        return None
    try:
        async with httpx.AsyncClient(timeout=30, follow_redirects=True) as client:
            resp = await client.get(url)
            if resp.status_code == 200 and b"%PDF" in resp.content[:10]:
                return resp.content
            logger.warning(f"PDF fetch failed ({resp.status_code}): {url}")
            return None
    except Exception as exc:
        logger.warning(f"PDF download error: {exc} — {url}")
        return None


def extract_text(pdf_bytes: bytes) -> str:
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            parts = []
            for page in pdf.pages:
                text = page.extract_text() or ""
                parts.append(text)
            return "\n".join(parts)
    except Exception as exc:
        logger.warning(f"PDF text extraction error: {exc}")
        return ""


# ─────────────────────────────────────────────────────────────────────────────
# Keyword search
# ─────────────────────────────────────────────────────────────────────────────

def extract_dispozitiv(text: str) -> str:
    """Returnează doar textul din dispozitivul hotărârii (după DISPUN/DISPUNE)."""
    text_lower = text.lower()
    for marker in ["dispune:", "dispune\n", "dispun:", "dispun\n", "dispune ", "dispun "]:
        pos = text_lower.rfind(marker)  # rfind = ultima apariție (dispozitivul e la final)
        if pos != -1:
            return text[pos:]
    return text  # fallback: tot textul dacă nu se găsește marcatorul


def find_keyword_excerpts(text: str) -> list[dict]:
    """Caută cuvintele cheie doar în dispozitiv și returnează un singur excerpt combinat."""
    dispozitiv = extract_dispozitiv(text)
    dispozitiv_lower = dispozitiv.lower()

    found_keywords = []
    excerpts = []
    seen_positions = set()

    for kw in KEYWORDS:
        kw_lower = kw.lower()
        start = 0
        while True:
            pos = dispozitiv_lower.find(kw_lower, start)
            if pos == -1:
                break
            bucket = pos // (CONTEXT_CHARS // 2)
            if bucket not in seen_positions:
                seen_positions.add(bucket)
                excerpt_start = max(0, pos - CONTEXT_CHARS // 2)
                excerpt_end = min(len(dispozitiv), pos + len(kw) + CONTEXT_CHARS // 2)
                excerpt = dispozitiv[excerpt_start:excerpt_end].strip()
                excerpt = " ".join(excerpt.split())
                if kw not in found_keywords:
                    found_keywords.append(kw)
                excerpts.append(excerpt)
            start = pos + 1

    if not found_keywords:
        return []

    # Toate mențiunile combinate într-un singur rezultat (un singur rând per cauză)
    return [{
        "keyword": ", ".join(found_keywords),
        "excerpt": "\n\n".join(excerpts),
    }]


# ─────────────────────────────────────────────────────────────────────────────
# Per-case API (folosit de main.py)
# ─────────────────────────────────────────────────────────────────────────────

def load_keywords() -> dict:
    """Încarcă toate categoriile de keywords din config/keywords.json."""
    default = {
        "case_filter": KEYWORDS,
        "asset_keywords": [
            "autoturism", "automobil", "vehicul",
            "imobil", "apartament", "casă", "teren", "lot",
            "mijloace bănești", "numerar", "cont bancar", "depozit",
            "lei", "MDL", "euro", "EUR", "dolari", "USD",
            "valori mobiliare", "acțiuni",
        ],
        "decision_keywords": [
            "dispune", "hotărăște", "declară",
            "condamnat", "achitat", "încetat",
            "pedeapsă", "amendă", "privațiune de libertate",
            "termen de probă", "muncă neremunerată",
        ],
    }
    try:
        with open(_CONFIG_PATH, encoding="utf-8") as f:
            cfg = json.load(f)
        result = {key: cfg.get(key) or default[key] for key in default}
        logger.info(f"Keywords încărcate din {_CONFIG_PATH}: {len(result['case_filter'])} cuvinte filtrare")
        return result
    except Exception as exc:
        logger.warning(f"Nu s-a putut citi config/keywords.json: {exc} — se folosesc keywords implicite")
        return default


def case_matches_filter(d: dict, filter_kws: list[str]) -> bool:
    """Filtru rapid pe metadate (fără descărcare PDF) înainte de analiza completă."""
    haystack = f"{d.get('tematica_dosarului', '')} {d.get('denumirea_dosarului', '')}".lower()
    return any(kw.lower() in haystack for kw in filter_kws)


def _find_excerpts(text: str, kws: list[str], max_excerpts: int = 12) -> tuple[list[str], list[str]]:
    """Caută `kws` în `text` și returnează (cuvinte găsite, excerpte de context)."""
    text_lower = text.lower()
    found_keywords: list[str] = []
    excerpts: list[str] = []
    seen_positions: set[int] = set()

    for kw in kws:
        kw_lower = kw.lower()
        start = 0
        while True:
            pos = text_lower.find(kw_lower, start)
            if pos == -1:
                break
            bucket = pos // (CONTEXT_CHARS // 2)
            if bucket not in seen_positions:
                seen_positions.add(bucket)
                excerpt_start = max(0, pos - CONTEXT_CHARS // 2)
                excerpt_end = min(len(text), pos + len(kw) + CONTEXT_CHARS // 2)
                excerpts.append(" ".join(text[excerpt_start:excerpt_end].split()))
                if kw not in found_keywords:
                    found_keywords.append(kw)
            start = pos + 1

    return found_keywords[:max_excerpts], excerpts[:max_excerpts]


async def analyze_case(d: dict, keywords: dict) -> dict:
    """Descarcă și analizează PDF-ul unui dosar; returnează structura cerută de generate_court_report."""
    url = d.get("act_judecatoresc_url", "")
    result = {
        "dosar": d,
        "pdf_url": url,
        "pdf_disponibil": False,
        "text_extras": False,
        "pagini": 0,
        "keywords_gasite": [],
        "sectiuni_confiscare": [],
        "sectiuni_bunuri": [],
        "sectiuni_decizie": [],
        "eroare": "",
    }

    if not url:
        result["eroare"] = "Link PDF indisponibil"
        return result

    pdf_bytes = await download_pdf(url)
    if pdf_bytes is None:
        result["eroare"] = "PDF indisponibil"
        return result
    result["pdf_disponibil"] = True

    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            result["pagini"] = len(pdf.pages)
    except Exception as exc:
        logger.warning(f"Nu s-a putut determina numărul de pagini: {exc}")

    text = extract_text(pdf_bytes)
    if not text:
        result["eroare"] = "Text neextras din PDF"
        return result
    result["text_extras"] = True

    kw_confiscare, exc_confiscare = _find_excerpts(text, keywords.get("case_filter", []))
    _, exc_bunuri = _find_excerpts(text, keywords.get("asset_keywords", []))
    _, exc_decizie = _find_excerpts(text, keywords.get("decision_keywords", []))

    result["keywords_gasite"] = kw_confiscare
    result["sectiuni_confiscare"] = exc_confiscare
    result["sectiuni_bunuri"] = exc_bunuri
    result["sectiuni_decizie"] = exc_decizie
    return result


# ─────────────────────────────────────────────────────────────────────────────
# Main analysis
# ─────────────────────────────────────────────────────────────────────────────

async def analyze_decisions(decisions: list[dict]) -> list[dict]:
    rows = []
    total = len(decisions)
    for i, d in enumerate(decisions, 1):
        url = d.get("act_judecatoresc_url", "")
        logger.info(f"[{i}/{total}] Analizez: {d.get('numarul_dosarului', '?')} — {url or 'fără PDF'}")

        base = {
            "instanta": d.get("instanta_judecatoreasca", ""),
            "nr_dosar": d.get("numarul_dosarului", ""),
            "denumire": d.get("denumirea_dosarului", ""),
            "data_pronuntarii": d.get("data_pronuntarii", ""),
            "judecator": d.get("judecator", ""),
            "tematica": d.get("tematica_dosarului", ""),
            "pdf_url": url,
        }

        if not url:
            rows.append({**base, "keyword": "", "excerpt": "", "status": STATUS_UNAVAILABLE})
            continue

        pdf_bytes = await download_pdf(url)
        if pdf_bytes is None:
            rows.append({**base, "keyword": "", "excerpt": "", "status": STATUS_UNAVAILABLE})
            continue

        text = extract_text(pdf_bytes)
        if not text:
            rows.append({**base, "keyword": "", "excerpt": "", "status": STATUS_UNAVAILABLE})
            continue

        excerpts = find_keyword_excerpts(text)
        if excerpts:
            # Întotdeauna un singur rând per cauză
            rows.append({**base, "keyword": excerpts[0]["keyword"], "excerpt": excerpts[0]["excerpt"], "status": STATUS_MATCH})
        else:
            rows.append({**base, "keyword": "", "excerpt": "", "status": STATUS_NO_MATCH})

    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Excel generation
# ─────────────────────────────────────────────────────────────────────────────

def build_excel(rows: list[dict], target_date: date) -> bytes:
    wb = Workbook()

    # ── Sheet 1: Date ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Analiză Confiscare"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1F3864")
    match_fill = PatternFill("solid", fgColor="FFF2CC")
    unavail_fill = PatternFill("solid", fgColor="D9D9D9")

    headers = [
        "Instanța", "Nr. dosar", "Denumire dosar",
        "Data pronunțării", "Judecător", "Tematică",
        "Cuvânt cheie", "Paragraf relevant", "Link PDF", "Status",
    ]
    col_widths = [28, 22, 40, 16, 24, 36, 22, 80, 18, 22]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for r, row in enumerate(rows, 2):
        values = [
            row["instanta"], row["nr_dosar"], row["denumire"],
            row["data_pronuntarii"], row["judecator"], row["tematica"],
            row["keyword"], row["excerpt"], row["pdf_url"], row["status"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=(col == 8))

        # Hyperlink PDF
        if row["pdf_url"]:
            link_cell = ws.cell(row=r, column=9)
            link_cell.hyperlink = row["pdf_url"]
            link_cell.value = "PDF"
            link_cell.font = Font(color="0563C1", underline="single")

        # Color rows
        status = row["status"]
        if status == STATUS_MATCH:
            fill = match_fill
        elif status == STATUS_UNAVAILABLE:
            fill = unavail_fill
        else:
            fill = None

        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=r, column=col).fill = fill

        ws.row_dimensions[r].height = 60 if row["excerpt"] else 18

    # ── Sheet 2: Statistici ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Statistici")
    title_font = Font(bold=True, size=12)
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 15

    total = len(rows)
    # Deduplicated by dosar for statistics
    dosare = {}
    for row in rows:
        key = row["nr_dosar"]
        if key not in dosare:
            dosare[key] = row["status"]
        elif row["status"] == STATUS_MATCH:
            dosare[key] = STATUS_MATCH

    n_match = sum(1 for s in dosare.values() if s == STATUS_MATCH)
    n_no_match = sum(1 for s in dosare.values() if s == STATUS_NO_MATCH)
    n_unavail = sum(1 for s in dosare.values() if s == STATUS_UNAVAILABLE)
    n_dosare = len(dosare)

    stat_rows = [
        ("Raport analiză confiscare/sechestru", ""),
        (f"Data analizată: {_day_ro(target_date)}", ""),
        ("", ""),
        ("Total dosare analizate", n_dosare),
        ("Cu mențiuni confiscare/sechestru", n_match),
        ("Fără mențiuni", n_no_match),
        ("PDF indisponibil", n_unavail),
        ("", ""),
        ("Total rânduri raport (incl. mențiuni multiple)", total),
    ]

    for r, (label, val) in enumerate(stat_rows, 1):
        ws2.cell(row=r, column=1, value=label)
        if val != "":
            ws2.cell(row=r, column=2, value=val)
        if r <= 2 or label.startswith("Total") or label.startswith("Cu"):
            ws2.cell(row=r, column=1).font = title_font

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Telegram
# ─────────────────────────────────────────────────────────────────────────────

async def send_message(token: str, chat_id: str, text: str) -> None:
    url = TELEGRAM_API_MSG.format(token=token)
    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.post(url, json={
            "chat_id": chat_id,
            "text": text,
            "parse_mode": "HTML",
            "disable_web_page_preview": True,
        })
        resp.raise_for_status()


async def send_document(token: str, chat_id: str, file_bytes: bytes, filename: str, caption: str) -> None:
    url = TELEGRAM_API_DOC.format(token=token)
    async with httpx.AsyncClient(timeout=60) as client:
        resp = await client.post(url, data={
            "chat_id": chat_id,
            "caption": caption,
            "parse_mode": "HTML",
        }, files={
            "document": (filename, file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        })
        if resp.status_code != 200:
            logger.error(f"Telegram sendDocument error {resp.status_code}: {resp.text}")
            resp.raise_for_status()


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

async def main() -> None:
    token = _require("TELEGRAM_BOT_TOKEN")
    chat_id = _require("TELEGRAM_CHAT_ID")

    target_date = date.today() - timedelta(days=1)
    logger.info(f"Target date: {target_date.strftime('%d.%m.%Y')}")

    decisions = await scrape_decisions(target_date)
    if not decisions:
        logger.info("Nu s-au găsit hotărâri — raport PDF nu este generat.")
        await send_message(token, chat_id,
            f"⚖️ <b>Analiză confiscare/sechestru</b>\n"
            f"📅 <b>{_day_ro(target_date)}</b>\n\n"
            "ℹ️ Nu au fost găsite hotărâri penale pentru această dată.")
        return

    logger.info(f"Analizez PDF-uri pentru {len(decisions)} hotărâri...")
    rows = await analyze_decisions(decisions)

    n_match = len({r["nr_dosar"] for r in rows if r["status"] == STATUS_MATCH})
    n_unavail = len({r["nr_dosar"] for r in rows if r["status"] == STATUS_UNAVAILABLE})
    logger.info(f"Rezultate: {n_match} dosare cu mențiuni, {n_unavail} PDF indisponibil")

    excel_bytes = build_excel(rows, target_date)

    filename = f"confiscare_{target_date.strftime('%Y-%m-%d')}.xlsx"
    with open(filename, "wb") as f:
        f.write(excel_bytes)
    logger.info(f"Excel salvat local: {filename}")

    caption = (
        f"⚖️ <b>Analiză confiscare/sechestru</b>\n"
        f"📅 <b>{_day_ro(target_date)}</b>\n\n"
        f"📊 Total hotărâri: <b>{len(decisions)}</b>\n"
        f"🔴 Cu mențiuni confiscare/sechestru: <b>{n_match}</b>\n"
        f"⚠️ PDF indisponibil: <b>{n_unavail}</b>"
    )
    await send_document(token, chat_id, excel_bytes, filename, caption)
    logger.info("Excel trimis pe Telegram.")


if __name__ == "__main__":
    asyncio.run(main())
